#Mukul
#2001CE35

#CAlling all required lib.

from platform import python_version
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import glob
import openpyxl
import os
from datetime import datetime
STRT_time = datetime.now()
os.system("cls")


ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename.

OCT_SIGN = [1, -1, 2, -2, 3, -3, 4, -4]
OCT_NAME_id_mapping = {1: "Internal outward interaction", -1: "External outward interaction", 2: "External Ejection", -
                       2: "Internal Ejection", 3: "External inward interaction", -3: "Internal inward interaction", 4: "Internal sweep", -4: "External sweep"}
YELLOW = "00FFFF00"
YELLOW_bg = PatternFill(
	start_color=YELLOW, end_color=YELLOW, fill_type='solid')
BLC = "00000000"
DoUbLe = Side(border_style="thin", color=BLC)
BLC_border = Border(top=DoUbLe, left=DoUbLe, right=DoUbLe, bottom=DoUbLe)

#Code


def RST_CNT(count):
    for item in OCT_SIGN:
        count[item] = 0

# Method to initialise dictionary with 0 for "OCT_SIGN" except 'left'


def RST_CNT_except(count, left):
    for item in OCT_SIGN:
        if (item != left):
            count[item] = 0


def SET_FREQ(longest, frequency, OUTsheet):
    # Iterating "OCT_SIGN" and updating sheet
    for i in range(9):
        for j in range(3):
            OUTsheet.cell(row=3+i, column=45+j).border = BLC_border

    OUTsheet.cell(row=3, column=45).value = "Octant ##"
    OUTsheet.cell(row=3, column=46).value = "Longest Subsquence Length"
    OUTsheet.cell(row=3, column=47).value = "Count"

    for i, label in enumerate(OCT_SIGN):
        currRow = i+3
        try:
            OUTsheet.cell(row=currRow+1, column=45).value = label
            OUTsheet.cell(column=46, row=currRow+1).value = longest[label]
            OUTsheet.cell(column=47, row=currRow+1).value = frequency[label]
        except FileNotFoundError:
            print("File not found!!")
            exit()

# Method to set time range for longest subsequence


def longest_subsequence_time(longest, frequency, timeRange, OUTsheet):
    # Naming columns number
    lengthCol = 50
    freqCol = 51

    # Initial row, just after the header row
    row = 4

    OUTsheet.cell(row=3, column=49).value = "Octant ###"
    OUTsheet.cell(row=3, column=50).value = "Longest Subsquence Length"
    OUTsheet.cell(row=3, column=51).value = "Count"

    # Iterating all octants
    for octant in OCT_SIGN:
        try:
            # Setting octant's longest subsequence and frequency data
            OUTsheet.cell(column=49, row=row).value = octant
            OUTsheet.cell(column=lengthCol, row=row).value = longest[octant]
            OUTsheet.cell(column=freqCol, row=row).value = frequency[octant]
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row += 1

        try:
            # Setting default labels
            OUTsheet.cell(column=49, row=row).value = "Time"
            OUTsheet.cell(column=lengthCol, row=row).value = "From"
            OUTsheet.cell(column=freqCol, row=row).value = "To"
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row += 1

        # Iterating time range values for each octants
        for timeData in timeRange[octant]:
            try:
                # Setting time interval value
                OUTsheet.cell(row=row, column=lengthCol).value = timeData[0]
                OUTsheet.cell(row=row, column=freqCol).value = timeData[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            OUTsheet.cell(row=i, column=j).border = BLC_border


def count_longest_subsequence_freq_func(longest, OUTsheet, total_count):
    # Dictionary to store consecutive sequence count
    count = {}

    # Dictionary to store frequency count
    frequency = {}

    # Dictionary to store time range
    timeRange = {}

    for label in OCT_SIGN:
        timeRange[label] = []

    # Initialing dictionary to 0 for all labels
    RST_CNT(count)
    RST_CNT(frequency)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, total_count):
        currRow = i+3
        try:
            curr = int(OUTsheet.cell(column=11, row=currRow).value)

            # Comparing current and last value
            if (curr == last):
                count[curr] += 1
            else:
                count[curr] = 1
                RST_CNT_except(count, curr)

            # Updating frequency
            if (count[curr] == longest[curr]):
                frequency[curr] += 1

                # Counting STRTing and ending time of longest subsequence
                end = float(OUTsheet.cell(row=currRow, column=1).value)
                STRT = 100*end - longest[curr]+1
                STRT /= 100

                # Inserting time interval into map
                timeRange[curr].append([STRT, end])

                # Resetting count dictionary
                RST_CNT(count)
            else:
                RST_CNT_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()

        # Updating 'last' variable
        last = curr

    # Setting frequency table into sheet
    SET_FREQ(longest, frequency, OUTsheet)

    # Setting time range for longest subsequence
    longest_subsequence_time(longest, frequency, timeRange, OUTsheet)

# Method to set frequency count to sheet


def find_longest_subsequence(OUTsheet, total_count):
	# Dictionary to store consecutive sequence count
    count = {}

    # Dictionary to store longest count
    longest = {}

    # Initialing dictionary to 0 for all labels
    RST_CNT(count)
    RST_CNT(longest)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, total_count):
        currRow = i+3
        try:
            curr = int(OUTsheet.cell(column=11, row=currRow).value)

            # Comparing current and last value
            if (curr == last):
                count[curr] += 1
                longest[curr] = max(longest[curr], count[curr])
                RST_CNT_except(count, curr)
            else:
                count[curr] = 1
                longest[curr] = max(longest[curr], count[curr])
                RST_CNT_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        # Updating "last" variable
        last = curr

    # Method to Count longest subsequence frequency
    count_longest_subsequence_freq_func(longest, OUTsheet, total_count)


def transition_count_func(row, transition_count, OUTsheet):
    # Setting hard coded inputs
    try:
        OUTsheet.cell(row=row, column=36).value = "To"
        OUTsheet.cell(row=row+1, column=35).value = "Octant #"
        OUTsheet.cell(row=row+2, column=34).value = "From"

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                OUTsheet.cell(row=j, column=i).border = BLC_border

    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    # Setting Labels
    for i, label in enumerate(OCT_SIGN):
        try:
            OUTsheet.cell(row=row+1, column=i+36).value = label
            OUTsheet.cell(row=row+i+2, column=35).value = label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    # Setting data
    for i, l1 in enumerate(OCT_SIGN):
        maxi = -1

        for j, l2 in enumerate(OCT_SIGN):
            val = transition_count[str(l1)+str(l2)]
            maxi = max(maxi, val)

        for j, l2 in enumerate(OCT_SIGN):
            try:
                OUTsheet.cell(row=row+i+2, column=36 +
                              j).value = transition_count[str(l1)+str(l2)]
                if transition_count[str(l1)+str(l2)] == maxi:
                    maxi = -1
                    OUTsheet.cell(row=row+i+2, column=36+j).fill = YELLOW_bg
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()


def set_mod_overall_transition_count(OUTsheet, mod, total_count):
	# Counting partitions w.r.t. mod
    try:
        totalPartition = total_count//mod
    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    # Checking mod value range
    if (mod < 0):
        raise Exception("Mod value should be in range of 1-30000")

    if (total_count % mod != 0):
        totalPartition += 1

    # Initializing row STRT for data filling
    rowSTRT = 16

    # Iterating all partitions
    for i in range(0, totalPartition):
        # Initializing STRT and end values
        STRT = i*mod
        end = min((i+1)*mod-1, total_count-1)

        # Setting STRT-end values
        try:
            OUTsheet.cell(column=35, row=rowSTRT-1 + 13 *
                          i).value = "Mod Transition Count"
            OUTsheet.cell(column=35, row=rowSTRT + 13 *
                          i).value = str(STRT) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Initializing empty dictionary
        transCount = {}
        for a in range(1, 5):
            for b in range(1, 5):
                transCount[str(a)+str(b)] = 0
                transCount[str(a)+str(-b)] = 0
                transCount[str(-a)+str(b)] = 0
                transCount[str(-a)+str(-b)] = 0

        # Counting transition for range [STRT, end)
        for a in range(STRT, end+1):
            try:
                curr = OUTsheet.cell(column=11, row=a+3).value
                next = OUTsheet.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            # Incrementing count for within range value
            if (next != None):
                transCount[str(curr) + str(next)] += 1

        # Setting transition counts
        transition_count_func(rowSTRT + 13*i, transCount, OUTsheet)


def set_overall_Transition_Count(OUTsheet, total_count):
	# Initializing empty dictionary
    count_transition = {}
    for i in range(1, 5):
        for j in range(1, 5):
            count_transition[str(i)+str(j)] = 0
            count_transition[str(i)+str(-j)] = 0
            count_transition[str(-i)+str(j)] = 0
            count_transition[str(-i)+str(-j)] = 0

    # Iterating octants values to fill dictionary
    STRT = 0

    # try and except block for string to int conversion
    try:
        last = int(OUTsheet["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()

    while (STRT < total_count-1):
        # try and except block for string to int conversion
        try:
            curr = int(OUTsheet.cell(row=STRT+4, column=11).value)
            count_transition[str(last) + str(curr)] += 1
            last = curr
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        STRT += 1

    # Setting transitions counted into sheet
    transition_count_func(2, count_transition, OUTsheet)

def set_rank_count(row,countMap, OUTsheet):
    # Copying the count list to sort
    sortedCount = []
    count = []
    for label in OCT_SIGN:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break
    rank1Oct = -10

    for j in range(0,8):
        OUTsheet.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = OCT_SIGN[j]
            OUTsheet.cell(row = row, column=23+j).fill = YELLOW_bg    

    OUTsheet.cell(row=row , column=31).value = rank1Oct
    OUTsheet.cell(row=row , column=32).value = OCT_NAME_id_mapping[rank1Oct]

def overall_octant_rank_func(last_row, OUTsheet):
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    row =4
    while OUTsheet.cell(row=row, column=29).value is not None:
        oct = int(OUTsheet.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1

    for i in range(9):
        for j in range(3):
            row = last_row+2+i
            col = 29+j
            OUTsheet.cell(row=row, column = col).border = BLC_border

    OUTsheet.cell(column=29, row=last_row+2).value = "Octant ID"
    OUTsheet.cell(column=30, row=last_row+2).value = "Octant Name "
    OUTsheet.cell(column=31, row=last_row+2).value = "Count of Rank 1 Mod Values"

    for j, oct in enumerate(OCT_SIGN):
        OUTsheet.cell(column=29, row=last_row+3+j).value = oct
        OUTsheet.cell(column=30, row=last_row+3+j).value = OCT_NAME_id_mapping[oct]
        OUTsheet.cell(column=31, row=last_row+3+j).value = count[oct]

def set_mod_count(OUTsheet, mod, total_count):
	# Initializing empty dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # Variable to store last row
    last_row = -1

    # Iterating loop to set count dictionary
    STRT = 0
    while(STRT<total_count):
        try:
            count[int(OUTsheet.cell(row=STRT+3, column=11).value)] +=1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        STRT+=1
        try:
            if(STRT%mod==0):
                # Setting row data
                try:
                    row = 4 + STRT//mod
                    last_row = row
                    OUTsheet.cell(row=row, column=14).value = str(STRT-mod) + "-" + str(min(total_count, STRT-1))

                    for i, label in enumerate(OCT_SIGN):
                        OUTsheet.cell(row=row, column=15+i).value = count[label]

                    set_rank_count(row,count, OUTsheet)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                # Reset count values
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("Mod can't have 0 value")
            exit()
    try:
        if(STRT%mod!=0):
            # Setting row data
            try:
                row = 5 + STRT//mod
                last_row = row
                OUTsheet.cell(row=row, column=14).value = str(STRT-mod) + "-" + str(min(total_count, STRT-1))
                for i, label in enumerate(OCT_SIGN):
                    OUTsheet.cell(row=row, column=15+i).value = count[label]
                
                set_rank_count(row,count, OUTsheet)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    if(last_row!=-1):
        overall_octant_rank_func(last_row, OUTsheet)

def setOverallCount(total_count, OUTsheet):	
	# Initializing count dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    # Incrementing count dictionary data
    try:
        for i in range (3,total_count+3):
            count[int(OUTsheet.cell(column=11, row=i).value)] = count[int(OUTsheet.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    # Setting data into sheet
    for i, label in enumerate(OCT_SIGN):
        try:
            OUTsheet.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    set_rank_count(4, count, OUTsheet)

def set_overall_octant_rank_count(OUTsheet, mod, total_count):
    headers = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    totalRows = total_count//mod+1+1 # header + overall
    if total_count%mod!=0:
        totalRows+=1

    for i, header in enumerate(headers):
        for j in range(totalRows):
            OUTsheet.cell(row=3+j, column = 14+i).border = BLC_border

    for i, header in enumerate(headers):
        OUTsheet.cell(row=3, column = i+14).value = header

    OUTsheet.cell(row=4, column = 13).value = "Mod " + str(mod)

    setOverallCount(total_count, OUTsheet)

# Method based on if-else to return octant type
def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2

    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3

    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4

def setProcessedDataWithOctant(U_AVG, V_AVG, W_AVG, total_count, inputSheet, OUTsheet):
    STRT = 2
    time = inputSheet.cell(STRT, 1).value

    # Iterating through out the sheet
    while(time!=None):
        # Calculating processed data
        try:
            U1 = inputSheet.cell(STRT, 2).value - U_AVG
            V1 = inputSheet.cell(STRT, 3).value - V_AVG
            W1 = inputSheet.cell(STRT, 4).value - W_AVG
            
            U1 = round(U1,3)
            V1 = round(V1,3)
            W1 = round(W1,3)

            oct = get_octant(U1, V1, W1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Setting processed data
        try:
            OUTsheet.cell(row=STRT+1, column=8).value = U1
            OUTsheet.cell(row=STRT+1, column=9).value = V1
            OUTsheet.cell(row=STRT+1, column=10).value = W1
            OUTsheet.cell(row=STRT+1, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        STRT = STRT+1
        try:
            time = inputSheet.cell(STRT, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

def SET_IN_DATA(input_file_name, OUTsheet):
	input_file = openpyxl.load_workbook(input_file_name)
	inputSheet = input_file.active

	STRT = 2
	time = inputSheet.cell(STRT, 1).value

    # Variables to store sum variable
	U_SUM = 0 
	V_SUM = 0
	W_SUM = 0

	# Iterating complete file till time value is not None
	while(time!=None):
		try:
			U_SUM += float(inputSheet.cell(STRT, 2).value)
			V_SUM += float(inputSheet.cell(STRT, 3).value)
			W_SUM += float(inputSheet.cell(STRT, 4).value)
		except ValueError:
			print("Sheet input can't be converted to float!!")
			exit()
		except TypeError:
			print("Sheet doesn't contain valid float input!!")
			exit()

		try:
			# Setting input time,u,v,w values
			OUTsheet.cell(row=STRT+1, column=1).value = inputSheet.cell(STRT, 1).value 
			OUTsheet.cell(row=STRT+1, column=2).value = inputSheet.cell(STRT, 2).value 
			OUTsheet.cell(row=STRT+1, column=3).value = inputSheet.cell(STRT, 3).value 
			OUTsheet.cell(row=STRT+1, column=4).value = inputSheet.cell(STRT, 4).value 
		except FileNotFoundError:
			print("File not found!!")
			exit()
		except ValueError:
			print("Row or column values must be at least 1 ")
			exit()

		STRT = STRT+1
		time = inputSheet.cell(STRT, 1).value

	# Setting total count
	total_count = STRT-2 # -1 for header and -1 for last None
	# Calculating average
	try:
		U_AVG = round(U_SUM/total_count, 3)
		V_AVG = round(V_SUM/total_count, 3)
		W_AVG = round(W_SUM/total_count, 3)
	except ZeroDivisionError:
		print("No input data found!!\nDivision by zero occurred!")
		exit()

	# Setting average values
	try:
		OUTsheet.cell(row=3, column=5).value = U_AVG
		OUTsheet.cell(row=3, column=6).value = V_AVG
		OUTsheet.cell(row=3, column=7).value = W_AVG
	except FileNotFoundError:
		print("Output file not found!!")
		exit()
	except ValueError:
		print("Row or column values must be at least 1 ")
		exit()

	# Processing input
	setProcessedDataWithOctant(U_AVG, V_AVG, W_AVG, total_count, inputSheet, OUTsheet)

	return total_count

def entry_point(input_file, mod):
	fileName = input_file.split("\\")[-1]
	fileName = fileName.split(".xlsx")[0]
	outputFileName = "output/" + fileName + "_octant_analysis_mod_" + str(mod) + ".xlsx"

	outputFile = openpyxl.Workbook()
	OUTsheet = outputFile.active

	OUTsheet.cell(row=1, column=14).value = "Overall Octant Count"
	OUTsheet.cell(row=1, column=24).value = "Rank #1 Should be highlighted YELLOW"
	OUTsheet.cell(row=1, column=35).value = "Overall Transition Count"
	OUTsheet.cell(row=1, column=45).value = "Longest Subsequence Length"
	OUTsheet.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
	OUTsheet.cell(row=2, column=36).value = "To"

	headers = ["T", "U", "V", "W", "U AVG", "V AVG", "W AVG", "U'=U - U AVG", "V'=V - V AVG","W'=W - W AVG", "Octant"]
	for i, header in enumerate(headers):
		OUTsheet.cell(row=2, column=i+1).value = header

	total_count = SET_IN_DATA(input_file, OUTsheet)
	set_overall_octant_rank_count(OUTsheet, mod, total_count)
	set_mod_count(OUTsheet, mod, total_count)
	set_overall_Transition_Count(OUTsheet, total_count)
	set_mod_overall_transition_count(OUTsheet, mod, total_count)
	find_longest_subsequence(OUTsheet, total_count)

	outputFile.save(outputFileName)

def octant_analysis(mod=5000):
	path = os.getcwd()
	csv_files = glob.glob(os.path.join(path + "\input", "*.xlsx"))
	
	for file in csv_files:
		entry_point(file, mod)
mod=5000
octant_analysis(mod)
