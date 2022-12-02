#Mukul: 2001CE35
#Divya: 2001CE60

#Importing all required libraries.
from datetime import datetime
start_time = datetime.now()

#Importing strimlit for front-end interface.
import streamlit as st
import requests #This will call files from web.
from streamlit_lottie import st_lottie #This will call lottie file(json file).
import pandas as pAnDa_jI
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from datetime import datetime

#This will be used for creating dialogue box(First part of code-strimlit.)
import tkinter as tk
from tkinter import filedialog

import glob
import os
from zipfile import ZipFile #This will be used for making zip folders.

#Already given part of the code.
from platform import python_version

os.system("cls")

ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")
 
#Here we will be reading all the files (or selected one) from input folder. (strictly .xlsx)
#After that code from tut.7 will be working in the backend which will chnange the files for a desired output and a proper naming scheme is followed.
#Finally all the files will be downlaoded or saved throgh the interface developed through streamlit.

#These are some formating criterias and values for color and other things which will be used in for formating the cells and tables of output.
OCT_SIGN = [1, -1, 2, -2, 3, -3, 4, -4]
OCT_NAME_ID_mapping = {1: "Internal outward interaction", -1: "External outward interaction", 2: "External Ejection", -2: "Internal Ejection", 3: "External inward interaction", -3: "Internal inward interaction", 4: "Internal sweep", -4: "External sweep"}
YELLOW = "00FFFF00"
YELLOW_bg = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
BLC = "00000000"
DoUbLe = Side(border_style="thin", color=BLC)
BLC_border = Border(top=DoUbLe, left=DoUbLe, right=DoUbLe, bottom=DoUbLe)

#This variable will be used to store the file uploaded on interface. 
new_file=None
MOD=0 #Initialised Mod value to 0.

#Function for front-end part.
def Streamlit_Proj2():

    #Setting the webpage title.
    st.set_page_config(page_title="Team: Mukul n Divya", page_icon=":hamster:", layout="wide")

    #Creating a container, it contains greetings n header.
    with st.container():
        st.subheader("Hi, :wave:")
        st.subheader("This is team Mukul n Divya")
        
    #This function will load the json file.
    def load_lottieurl(url):
        r = requests.get(url)
        if r.status_code != 200:
            return None
        return r.json()

    #Calling the json file(kinda gif) to program.
    lottie_coding = load_lottieurl("https://assets5.lottiefiles.com/packages/lf20_fcfjwiyb.json")

    path="" #This will be storing the path of folder from where we will be uploading our files.
    
    #This container have two columns, one with the main interface, other with gif.
    with st.container():
        #Setting the header value of the webpage
        left_column, right_column = st.columns(2)
        with left_column:
            st.subheader("Welcome to the Front-End Interface for Project 2.")
            
            #Using the radio button, to give options for the type of conversion.
            status = st.radio("Select conversion type: ", ('Single file Conversion', 'Bulk file Conversion'))
            
            #On selecting Single file Conversion.
            if (status == 'Single file Conversion'):
                global new_file
                
                #Using file_uploader for uploading the file.
                new_file = st.file_uploader("Upload the input dataset: ", type="xlsx")     #Strictly for .xlsx files only.
                if "path" in st.session_state: #Basically for re-running the program.
                    del st.session_state["path"]
            
            #On selecting Bulk file Conversion.
            if (status=='Bulk file Conversion'):
                #Here we will be starting by selecting the folder.
                new_file=None
                
                #This will be used for bringing the dialogue box, where option for selecting input folder will be given.
                root = tk.Tk()
                root.withdraw()
                root.wm_attributes('-topmost', 1) #This will open the dialog box in top left part of screen.

                #Creating a button to pick folder.
                st.write('Please select a folder:')
                clicked = st.button('Choose the folder')
                if clicked:
                    path = filedialog.askdirectory(master=root) #Through this the dialogue box will ask for directory.
                    st.session_state["path"] = path

            #Brings back to gui window.
            if "path" in st.session_state:
                path = st.session_state["path"]
                dirname = st.text_input('Selected folder:', path) #Here the addresss of that folder is shown.

        #This Column contains lottie file.
        with right_column:
            st_lottie(lottie_coding, height=300, key="coding")
        
        global MOD
        #After selecting the input files, here we will collect the mod value.
        MOD = st.number_input('Enter MOD value: ', min_value=1,  value=5000, step=1)
        
        #Devided the section in 2 columns, left side will have button for conversion, whereas right side will have button for downloading the file.
        convert, download=st.columns(2)
        #Download(Wanna Download) button will apppear only if, convert(Compute) button is used.
        with convert:
            conv_but=st.button("Compute") #Button for conversion.
                
            #After clicling on Compute.
            if conv_but:
                #Case for Single file Conversion.
                if (status == 'Single file Conversion'): 
                    
                    #To assure that a file is selected.
                    if not new_file:
                        st.warning("Please upload a file!!")
                    else:
                            
                        #Removing file extension cause we will be passing only file name to the main function(tut 7).
                        #Saving the nam in File_Name.
                        File_Name=new_file.name.split(".xlsx")[0]
                        outputFileName=proj2(File_Name) #Passing this name to main function.
                        
                        with download:
                            #Opening/recieving the output file, and dowloading it.
                            with open(outputFileName, 'rb') as req_file:
                                st.download_button(label="Wanna Download", data=req_file, file_name=outputFileName, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                                
                                
                    
                #Case of Bulk file conversion.      
                elif (status=='Bulk file Conversion'):
                    if "path" in st.session_state:
                        path = st.session_state["path"]

                    #Here we will be using the path of the folder and using a loop we will convert all the files.
                    folder = path.split("/")[-1]
                    excel_files=glob.glob(os.path.join(path,"*.xlsx"))#All the files in input with .xlsx extension.
                        
                    #To assure that required input is selected.
                    if(len(folder)==0):
                        st.warning("Atleast select a folder!!")
                        return
                    if len(excel_files)==0:
                        st.warning("No excel file present!!")
                        return 
                        
                    #Using the give namimg scheme for namimg the output folder.
                    outputFolderName=getName(folder)+".zip"
                        
                    #ZipFile will zip the folder.
                    zipObj=ZipFile(outputFolderName,'w')
                        
                    #Using loop and moving through all the excel files prsent in input and converting them to output.
                    for i, file in enumerate(excel_files):
                        new_file=file
                        File_Name=file.split(".xlsx")[0]
                        File_Name=File_Name.split("\\")[-1]
                            
                        #Calling main function for conversion of each file of excel files. 
                        outputFileName=proj2(File_Name)
                        
                        #Adding the file to zip folder.
                        zipObj.write(outputFileName)
                            
                    zipObj.close()
                        
                    #Finally downloading the zip folder.
                    with download:
                        with open(outputFolderName,'rb') as req_file:
                            st.download_button(label="Wanna Downalod", data=req_file, file_name=outputFolderName)
                     
                            

#This function is used for implemention naming scheme.
def getName(inputFileName):
    current=datetime.now()
    #Adding date n time.
    dt_s=current.strftime("%Y-%m-%d-%H-%M-%S")
    
    #Adding output/ in front for saving all the files in output folder after changing the names.
    outputFileName="output/"+inputFileName+"_MOD_"+str(MOD)+"_"+dt_s
    return outputFileName

#This is a modified version of tut 7 code, although all the function used inside this are same.
def proj2(File_Name):
    DaTaFrAmE=pAnDa_jI.read_excel(new_file)
    outputFileName=getName(File_Name)+".xlsx"
    
    outputFile=openpyxl.Workbook()
    OUTsheet=outputFile.active
    
    outputFile = openpyxl.Workbook()
    OUTsheet = outputFile.active
    TOT_Count = 0

    col = 1

    #Declairing variables to store sum values.
    U_SUM = 0 
    V_SUM = 0
    W_SUM = 0

    for key, value in DaTaFrAmE.items():
        value = value.tolist()
        TOT_Count = len(value)

        #Key shifted to 2nd row.
        OUTsheet.cell(row=2, column=col).value = key

        for r, val in enumerate(value):
            if col==2:
                U_SUM += val
            elif col==3:
                V_SUM += val
            elif col==4:
                W_SUM += val

            OUTsheet.cell(row=r+3, column=col).value = val        
        col +=1

    #Calculating averages.
    try:
        U_AVG = round(U_SUM/TOT_Count, 3)
        V_AVG = round(V_SUM/TOT_Count, 3)
        W_AVG = round(W_SUM/TOT_Count, 3)
    except ZeroDivisionError:
        print("No input data found!!\nDivision by zero occurred!")
        exit()

    #Inserting average values in output sheet.
    try:
        OUTsheet.cell(row=3, column=5).value = U_AVG
        OUTsheet.cell(row=3, column=6).value = V_AVG
        OUTsheet.cell(row=3, column=7).value = W_AVG
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    #After getting input ready, here we will set the data with octants.
    Pairing_Processed_Data_n_Octants(U_AVG, V_AVG, W_AVG, OUTsheet)

    OUTsheet.cell(row=1, column=14).value = "Overall Octant Count"
    OUTsheet.cell(row=1, column=24).value = "Rank #1 Should be highlighted yellow"
    OUTsheet.cell(row=1, column=35).value = "Overall Transition Count"
    OUTsheet.cell(row=1, column=45).value = "Longest Subsequence Length"
    OUTsheet.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
    OUTsheet.cell(row=2, column=36).value = "To"

    headers = ["T", "U", "V", "W", "U AVG", "V AVG", "W AVG","U'=U - U AVG", "V'=V - V AVG", "W'=W - W AVG", "Octant"]
    for i, header in enumerate(headers):
        OUTsheet.cell(row=2, column=i+1).value = header


    Overall_Octant_Rank_Count(OUTsheet, MOD, TOT_Count)

    Set_MODwise_Count(OUTsheet, MOD, TOT_Count)

    Overall_Transition_Count(OUTsheet, TOT_Count)

    #Function to add MOD wise count of transition.
    MODwise_Overall_Transition_Count(OUTsheet, MOD, TOT_Count)

    Longest_SubSeq(OUTsheet, TOT_Count)

    outputFile.save(outputFileName)


    data = OUTsheet.values
    columns = next(data)[0:]


    DaTaFrAmE = pAnDa_jI.DataFrame(data, columns=columns)
    return outputFileName

#Setting count to 0 for octants.
def RST_CNT(count):
    for item in OCT_SIGN:
        count[item] = 0

#Method to initialise dictionary with 0 for "OCT_SIGN" except 'left'.
def RST_CNT_except(count, left):
    for item in OCT_SIGN:
        if(item!=left):
            count[item] = 0

def Longest_SubSeq(OUTsheet, TOT_Count):
	#This Dictinary will store consecutive sequence count.
    count = {}

    #This Dictionary wil store longest count.
    longest = {}

    #Initialing dictionary to 0 for all labels.
    RST_CNT(count)
    RST_CNT(longest)

    #Variable to check end value.
    end = -10

    #Iterating complete excel sheet.
    for i in range(0, TOT_Count):
        currRow = i+3
        try:
            curr = int(OUTsheet.cell(column=11, row=currRow).value)

            #Comparing current and end value.
            if(curr==end):
                count[curr]+=1
                longest[curr] = max(longest[curr], count[curr])
                RST_CNT_except(count, curr)
            else:
                count[curr]=1
                longest[curr] = max(longest[curr], count[curr])
                RST_CNT_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        #Updating "end" variable.
        end = curr

    #Method to Count longest subsequence freq.
    Longest_SubSeq_freq(longest, OUTsheet, TOT_Count)

def Longest_SubSeq_freq(longest, OUTsheet, TOT_Count):
    #Dictinary to store consecutive sequence count.
    count = {}

    #Dictinary to store freq count.
    freq = {}

    #Dictionary to store time range.
    timeRange = {}

    for label in OCT_SIGN:
        timeRange[label] = []

    #Initialing dictionary to 0 for all labels.
    RST_CNT(count)
    RST_CNT(freq)

    #Variable to check end value.
    end = -10

    #Iterating complete excel sheet.
    for i in range(0, TOT_Count):
        currRow = i+3
        try:
            curr = int(OUTsheet.cell(column=11, row=currRow).value)
            
            #Comparing current and end value.
            if(curr==end):
                count[curr]+=1
            else:
                count[curr]=1        
                RST_CNT_except(count, curr)

            #Upading freq.
            if(count[curr]==longest[curr]):
                freq[curr]+=1

                #Counting starting and ending time of longest subsequence.
                end = float(OUTsheet.cell(row=currRow, column=1).value)
                start = 100*end - longest[curr]+1
                start/=100

                #Inserting time interval into map.
                timeRange[curr].append([start, end])

                #Reseting count dictionary.
                RST_CNT(count)
            else:
                RST_CNT_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()
        #Updating 'end' variable.
        end = curr

    #Setting freq table into sheet.
    SET_FREQ(longest, freq, OUTsheet)

    #Setting time range for longest subsequence.
    Longest_SubSeq_time(longest, freq, timeRange, OUTsheet)

#Method to set freq count to sheet.
def SET_FREQ(longest, freq, OUTsheet):
    #Iterating "OCT_SIGN" and updating sheet.
    for i in range(9):
        for j in range(3):
            OUTsheet.cell(row = 3+i, column = 45+j).border = BLC_border

    OUTsheet.cell(row=3, column=45).value= "Octant ##"
    OUTsheet.cell(row=3, column=46).value= "Longest Subsquence Length"
    OUTsheet.cell(row=3, column=47).value= "Count"

    for i, label in enumerate(OCT_SIGN):
        currRow = i+3
        try:
            OUTsheet.cell(row=currRow+1, column=45).value = label	
            OUTsheet.cell(column=46, row=currRow+1).value = longest[label]
            OUTsheet.cell(column=47, row=currRow+1).value = freq[label]
        except FileNotFoundError:
            print("File not found!!")
            exit()

#Method to set time range for longest subsequence.
def Longest_SubSeq_time(longest, freq, timeRange, OUTsheet):
    #Naming columns number.
    lengthCol = 50
    freqCol = 51
    
    #Initial row, just after the header row.
    row = 4

    OUTsheet.cell(row=3, column = 49).value = "Octant ###"
    OUTsheet.cell(row=3, column = 50).value = "Longest Subsquence Length"
    OUTsheet.cell(row=3, column = 51).value = "Count"

    #Iterating all octants. 
    for octant in OCT_SIGN:
        try:
            #Setting octant's longest subsequence and freq data.
            OUTsheet.cell(column=49, row=row).value = octant
            OUTsheet.cell(column=lengthCol, row=row).value = longest[octant]
            OUTsheet.cell(column=freqCol, row=row).value = freq[octant]
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        try:
            #Setting default labels.
            OUTsheet.cell(column=49, row=row).value = "Time"
            OUTsheet.cell(column=lengthCol, row=row).value = "From"
            OUTsheet.cell(column=freqCol, row=row).value = "To"
        except FileNotFoundError:
            print("File not found!!")
            exit()
        row+=1

        #Iterating time range values for each octants.
        for timeData in timeRange[octant]:
            try:
                #Setting time interval value.
                OUTsheet.cell(row=row, column=lengthCol).value = timeData[0]
                OUTsheet.cell(row=row, column=freqCol).value = timeData[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            OUTsheet.cell(row=i, column = j).border = BLC_border

def MODwise_Overall_Transition_Count(OUTsheet, MOD, TOT_Count):
    #Counting partitions w.r.t. MOD.
    try:
        totalPartition = TOT_Count//MOD
    except ZeroDivisionError:
        print("MOD can't have 0 value")
        exit()

    #Checking MOD value range.
    if(MOD<0):
        raise Exception("MOD value should be in range of 1-30000")

    if(TOT_Count%MOD!=0):
        totalPartition +=1

    #Initialising row start for data filling.
    rowStart = 16

    #Iterating all partitions.
    for i in range (0,totalPartition):
        #Initialising start and end values.
        start = i*MOD
        end = min((i+1)*MOD-1, TOT_Count-1)

        #Setting start-end values.
        try:
            OUTsheet.cell(column=35, row=rowStart-1 + 13*i).value = "MOD Transition Count"
            OUTsheet.cell(column=35, row=rowStart + 13*i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        #Initialising empty dictionary.
        transCount = {}
        for a in range (1,5):
            for b in range(1,5):
                transCount[str(a)+str(b)]=0
                transCount[str(a)+str(-b)]=0
                transCount[str(-a)+str(b)]=0
                transCount[str(-a)+str(-b)]=0
                
        #Counting transition for range [start, end].
        for a in range(start, end+1):
            try:
                curr = OUTsheet.cell(column=11, row=a+3).value
                next = OUTsheet.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            #Incrementing count for within range value.
            if(next!=None):
                transCount[str(curr) + str(next)]+=1

        #Setting transition counts.
        setTransitionCount(rowStart + 13*i, transCount, OUTsheet)

def Overall_Transition_Count(OUTsheet, TOT_Count):
    #Setting value.

    #Initialising empty dictionary.
    transCount = {}
    for i in range (1,5):
        for j in range(1,5):
            transCount[str(i)+str(j)]=0
            transCount[str(i)+str(-j)]=0
            transCount[str(-i)+str(j)]=0
            transCount[str(-i)+str(-j)]=0
        
    #Iterating octants values to fill dictionary.
    start = 0

    #try and except block for string to int conversion.
    try:
        end = int(OUTsheet["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()

    while(start<TOT_Count-1):
        #try and except block for string to int conversion.
        try:
            curr = int(OUTsheet.cell(row= start+4, column=11).value)
            transCount[str(end) + str(curr)]+=1
            end = curr
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1
    
    #Setting transitions counted into sheet.
    setTransitionCount(2, transCount, OUTsheet)

#Function to set Transition count.
def setTransitionCount(row, transCount, OUTsheet):
    #Setting hard coded inputs.
    try:
        OUTsheet.cell(row=row, column=36).value = "To"
        OUTsheet.cell(row=row+1, column=35).value = "Octant #"
        OUTsheet.cell(row=row+2, column=34).value = "From"

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                OUTsheet.cell(row=j, column = i).border = BLC_border

    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    #Setting Labels.
    for i, label in enumerate(OCT_SIGN):
        try:
            OUTsheet.cell(row=row+1, column=i+36).value=label
            OUTsheet.cell(row=row+i+2, column=35).value=label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    #Setting data.
    for i, l1 in enumerate(OCT_SIGN):
        maxi = -1

        for j, l2 in enumerate(OCT_SIGN):
            val = transCount[str(l1)+str(l2)]
            maxi = max(maxi, val)

        for j, l2 in enumerate(OCT_SIGN):
            try:
                OUTsheet.cell(row=row+i+2, column=36+j).value = transCount[str(l1)+str(l2)]
                if transCount[str(l1)+str(l2)] == maxi:
                    maxi = -1
                    OUTsheet.cell(row=row+i+2, column=36+j).fill = YELLOW_bg
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

def Set_MODwise_Count(OUTsheet, MOD, TOT_Count):
	#Initialising empty dictionary.
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    #Variable to store end row.
    lastRow = -1

    #Iterating loop to set count dictionary.
    start = 0
    while(start<TOT_Count):
        try:
            count[int(OUTsheet.cell(row=start+3, column=11).value)] +=1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start+=1
    
        try:    
            if(start%MOD==0):
                #Setting row data.
                try:
                    row = 4 + start//MOD
                    lastRow = row
                    OUTsheet.cell(row=row, column=14).value = str(start-MOD) + "-" + str(min(TOT_Count, start-1))

                    for i, label in enumerate(OCT_SIGN):
                        OUTsheet.cell(row=row, column=15+i).value = count[label]

                    SET_RANK_Count(row,count, OUTsheet)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                #Reset count values.
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("MOD can't have 0 value")
            exit()

    try:
        if(start%MOD!=0):
            #Setting row data.
            try:
                row = 5 + start//MOD
                lastRow = row
                OUTsheet.cell(row=row, column=14).value = str(start-MOD) + "-" + str(min(TOT_Count, start-1))
                for i, label in enumerate(OCT_SIGN):
                    OUTsheet.cell(row=row, column=15+i).value = count[label]
                
                SET_RANK_Count(row,count, OUTsheet)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("MOD can't have 0 value")
        exit()

    if(lastRow!=-1):
        setOverallOctantRankMap(lastRow, OUTsheet)

def setOverallOctantRankMap(lastRow, OUTsheet):
    count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    
    row =4
    while OUTsheet.cell(row=row, column=29).value is not None:
        oct = int(OUTsheet.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1

    for i in range(9):
        for j in range(3):
            row = lastRow+2+i
            col = 29+j
            OUTsheet.cell(row=row, column = col).border = BLC_border

    OUTsheet.cell(column=29, row=lastRow+2).value = "Octant ID"
    OUTsheet.cell(column=30, row=lastRow+2).value = "Octant Name "
    OUTsheet.cell(column=31, row=lastRow+2).value = "Count of Rank 1 MOD Values"

    for j, oct in enumerate(OCT_SIGN):
        OUTsheet.cell(column=29, row=lastRow+3+j).value = oct
        OUTsheet.cell(column=30, row=lastRow+3+j).value = OCT_NAME_ID_mapping[oct]
        OUTsheet.cell(column=31, row=lastRow+3+j).value = count[oct]

def Overall_Octant_Rank_Count(OUTsheet, MOD, TOT_Count):
    firstRow = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    totalRows = TOT_Count//MOD+1+1 
    if TOT_Count%MOD!=0:
        totalRows+=1

    for i, header in enumerate(firstRow):
        for j in range(totalRows):
            OUTsheet.cell(row=3+j, column = 14+i).border = BLC_border

    for i, header in enumerate(firstRow):
        OUTsheet.cell(row=3, column = i+14).value = header

    OUTsheet.cell(row=4, column = 13).value = "MOD " + str(MOD)
    setOverallCount(TOT_Count, OUTsheet)

def setOverallCount(TOT_Count, OUTsheet):	
	#Initialising count dictionary.
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    #Incrementing count dictionary data.
    try:
        for i in range (3,TOT_Count+3):
            count[int(OUTsheet.cell(column=11, row=i).value)] = count[int(OUTsheet.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    #Setting data into sheet.
    for i, label in enumerate(OCT_SIGN):
        try:
            OUTsheet.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()
    SET_RANK_Count(4, count, OUTsheet)

def SET_RANK_Count(row,countMap, OUTsheet):
    #Copying the count list to sort.
    sortedCount = []
    count = []
    for label in OCT_SIGN:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)
    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break

    rank1Oct = -10

    for j in range(0,8):
        OUTsheet.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = OCT_SIGN[j]
            OUTsheet.cell(row = row, column=23+j).fill = YELLOW_bg   

    OUTsheet.cell(row=row , column=31).value = rank1Oct
    OUTsheet.cell(row=row , column=32).value = OCT_NAME_ID_mapping[rank1Oct]

def Pairing_Processed_Data_n_Octants(U_AVG, V_AVG, W_AVG, OUTsheet):
    start = 3
    time = OUTsheet.cell(start, 1).value

    #Iterating throught sheet.
    while(time!=None):
        #Calculating processed data.
        try:
            u1 = OUTsheet.cell(start, 2).value - U_AVG
            v1 = OUTsheet.cell(start, 3).value - V_AVG
            w1 = OUTsheet.cell(start, 4).value - W_AVG
            
            u1 = round(u1,3)
            v1 = round(v1,3)
            w1 = round(w1,3)

            oct = get_octant(u1, v1, w1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        #Setting processed data.
        try:
            OUTsheet.cell(row=start, column=8).value = u1
            OUTsheet.cell(row=start, column=9).value = v1
            OUTsheet.cell(row=start, column=10).value = w1
            OUTsheet.cell(row=start, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start = start+1
        try:
            time = OUTsheet.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

#Method based on if-else to return octant type.
def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2
    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3
    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4
        
#Finally calling fuction for interface.   
Streamlit_Proj2()
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))