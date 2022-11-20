#Mukul 2001CE35

#Calling all required libraries.
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border,Side
import numpy as NUM_py
import os
import re
from datetime import datetime
os.system('cls')
start_time=datetime.now()

#This function will work only for one inning.
#We will call it again for the second inning.
def INNING(inn1,bat_pl,bow_pl,s): 
    innbat=Workbook()
    innfow=Workbook()
    innbow=Workbook()
    SHEET1=innbat.active
    SHEET2=innbow.active
    SHEET3=innfow.active
    SHEET1.column_dimensions['A'].width=25

    #Creating column headings for scorecard and index.
    SHEET1['A1']=s+' Innings'
    SHEET1['I1']='0-0'
    SHEET1['J1']='0 overs'
    SHEET1['A2']='Batter'
    SHEET1['F2']='R'
    SHEET1['G2']='B'
    SHEET1['H2']='4s'
    SHEET1['I2']='6s'
    SHEET1['J2']='SR'
    SHEET2['A1']='Bowler'
    SHEET2['D1']='O'
    SHEET2['E1']='M'
    SHEET2['F1']='R'
    SHEET2['G1']='W'
    SHEET2['H1']='NB'
    SHEET2['I1']='WD'
    SHEET2['J1']='ECO'
    SHEET3['A1']='Fall of wickets'
    
    #Using regex.
    
    #These are some variables which will be storing the data which relates to their name.
    over=re.compile(r'(\d\d?\.\d)')
    zero=re.compile(r'no run')
    no_ball=re.compile(r', (no ball),')
    wide=re.compile(r', wide,')
    wide2=re.compile(r', 2 wides,')
    wide3=re.compile(r', 3 wides,')
    single=re.compile(r', 1 run,')
    six=re.compile(r', SIX,')
    four=re.compile(r', FOUR,')
    byes=re.compile(r', (byes),')
    lbyes=re.compile(r', (leg byes),')
    double=re.compile(r', 2 runs,')
    triple=re.compile(r', 3 runs,')
    out=re.compile(r', out')
    player=re.compile(r'(\d\d?\.\d) (\w+) (to|\w+ to) (\w+)( \w+)?,')
    caught=re.compile(r', out Caught by (\w+)')
    lbw=re.compile(r', out Lbw!!')
    bowled=re.compile(r', out Bowled!!')
    run_out=re.compile(r'Run Out!! ')
    runs=0
    wickets=0
    nb=0
    nlb=0
    nw=0
    nnb=0
    ppr=0
    
    #This code dynamically works on each line.
    while True:
        t = inn1.readline()
        if not t:
            break
        crun = 0
        cex = 0
        ov = over.findall(t)
        if float(ov[0]) == int(float(ov[0]))+0.6:
            ov[0] = str(int(float(ov[0]))+1)
        if float(ov[0]) == int(float(ov[0]))+0.1:
            orun = runs
        nnnb = no_ball.findall(t)
        wbb = wide.findall(t)
        sr = single.findall(t)
        zr = zero.findall(t)
        by = byes.findall(t)
        lby = lbyes.findall(t)
        pl = player.finditer(t)
        w2 = wide2.findall(t)
        w3 = wide3.findall(t)
        for i in pl:
            cbow = i.group(2)
            cbat = i.group(4)
        r = 1
        for col in SHEET1.iter_cols(min_col=1, max_col=1):
            for cell in col:
                if cell.value == bat_pl[cbat]:
                    crow = cell.row
                    r = 0
                    break
        if r:
            SHEET1.append([bat_pl[cbat], 'not out', '', '', '', 0, 0, 0, 0, 0])
            crow = SHEET1.max_row

        s = 1
        for col in SHEET2.iter_cols(min_col=1, max_col=1):
            for cell in col:
                if cell.value == bow_pl[cbow]:
                    cbrow = cell.row
                    s = 0
                    break
        if s:
            SHEET2.append([bow_pl[cbow], '', '', 0, 0, 0, 0, 0, 0, 0])
            cbrow = SHEET2.max_row

		#Using if else statements for different rules of cricket.
        if sr:
            crun = 1
        elif zr:
            pass
        else:
            db = double.findall(t)
            if db:
                crun = 2
            else:
                fr = four.findall(t)
                if fr:
                    crun = 4
                    if not (by or lby):
                        SHEET1.cell(row=crow, column=8).value = SHEET1.cell(
                            row=crow, column=8).value+1
                else:
                    sx = six.findall(t)
                    if sx:
                        crun = 6
                        SHEET1.cell(row=crow, column=9).value = SHEET1.cell(
                            row=crow, column=9).value+1
                    else:
                        tp = triple.findall(t)
                        if tp:
                            crun = 3
        if wbb or nnnb or w2 or w3:
            cex = 1
            SHEET2.cell(row=cbrow, column=6).value = SHEET2.cell(
                row=cbrow, column=6).value+1
            if wbb:
                nw = nw+1
                SHEET2.cell(row=cbrow, column=9).value = SHEET2.cell(
                    row=cbrow, column=9).value+1
            elif w2:
                nw = nw+2
                cex = 2
                SHEET2.cell(row=cbrow, column=6).value = SHEET2.cell(
                    row=cbrow, column=6).value+1
                SHEET2.cell(row=cbrow, column=9).value = SHEET2.cell(
                    row=cbrow, column=9).value+2
            elif w3:
                nw = nw+3
                cex = 3
                SHEET2.cell(row=cbrow, column=6).value = SHEET2.cell(
                    row=cbrow, column=6).value+2
                SHEET2.cell(row=cbrow, column=9).value = SHEET2.cell(
                    row=cbrow, column=9).value+3
            else:
                nnb = nnb+1
                SHEET1.cell(row=crow, column=7).value = SHEET1.cell(
                    row=crow, column=7).value+1
                SHEET2.cell(row=cbrow, column=8).value = SHEET2.cell(
                    row=cbrow, column=8).value+1
        else:
            SHEET1.cell(row=crow, column=7).value = SHEET1.cell(
                row=crow, column=7).value+1
            bo = 10*SHEET2.cell(row=cbrow, column=4).value - \
                int(SHEET2.cell(row=cbrow, column=4).value)*4+1
            SHEET2.cell(row=cbrow, column=4).value = int(bo/6)*0.4+bo*0.1
        runs = runs+crun+cex
        if float(ov[0]) < 6.1:
            ppr = runs
        if float(ov[0]) == int(float(ov[0])):
            if orun == runs:
                SHEET2.cell(row=cbrow, column=5).value = SHEET2.cell(
                    row=cbrow, column=5).value+1
        ot = out.findall(t)
        ct = caught.findall(t)
        lw = lbw.findall(t)
        bw = bowled.findall(t)
        ctu = caught.finditer(t)
        ro = run_out.findall(t)
        for i in ctu:
            capl = i.group(1)
        if ot:
            wickets = wickets+1
            if not ro:
                SHEET2.cell(row=cbrow, column=7).value = SHEET2.cell(
                    row=cbrow, column=7).value+1
            if wickets != 1:
                SHEET3['A2'] = str(SHEET3['A2'].value)+', '+str(runs) + \
                    '-'+str(wickets)+' ('+bat_pl[cbat]+', '+ov[0]+')'
            else:
                SHEET3['A2'] = str(runs)+'-'+str(wickets) + \
                    '('+bat_pl[cbat]+', '+ov[0]+')'
            if ct:
                SHEET1.cell(row=crow, column=2).value = 'c ' + \
                    bow_pl[capl]+' b '+bow_pl[cbow]
            elif lw:
                SHEET1.cell(row=crow, column=2).value = 'lbw b '+bow_pl[cbow]
            elif bw:
                SHEET1.cell(row=crow, column=2).value = 'b '+bow_pl[cbow]
            elif ro:
                SHEET1.cell(
                    row=crow, column=2).value = 'run out ('+bow_pl[cbow]+')'

		#Now filling this data in exel file.
        SHEET1['I1'] = str(runs)+'-'+str(wickets)+'('+ov[0]+' Ov)'

        if not (by or lby):
            SHEET1.cell(row=crow, column=6).value = SHEET1.cell(
                row=crow, column=6).value+crun
            SHEET2.cell(row=cbrow, column=6).value = SHEET2.cell(
                row=cbrow, column=6).value+crun
        else:
            if by:
                if fr:
                    nb = nb+4
                elif db:
                    nb = nb+2
                elif tp:
                    nb = nb+3
                elif sr:
                    nb = nb+1
            else:
                if sr:
                    nlb = nlb+1
                elif fr:
                    nlb = nlb+4
                elif db:
                    nlb = nlb+2
                elif tp:
                    nlb = nlb+3

        SHEET1.cell(row=crow, column=10).value = float("{:.2f}".format(
            (SHEET1.cell(row=crow, column=6).value)*100/SHEET1.cell(row=crow, column=7).value))
        ovf = (10*SHEET2.cell(row=cbrow, column=4).value - 4 *
               int(SHEET2.cell(row=cbrow, column=4).value))/6
        if ovf:
            SHEET2.cell(row=cbrow, column=10).value = float(
                "{:.2f}".format(SHEET2.cell(row=cbrow, column=6).value/ovf))
        t = inn1.readline()
        if not t:
            break
        
    #Formatting the cells for a better display of info.
    SHEET1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    SHEET1.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
    crow = SHEET1.max_row+1
    SHEET1.merge_cells(start_row=crow, start_column=5,
                       end_row=crow, end_column=7)

    SHEET1.append(['Extras', '', '', '', '', '', '', str(nb+nlb+nw+nnb) +
                  '(b '+str(nb)+', lb '+str(nlb)+', w '+str(nw)+', nb '+str(nnb)+', p 0)'])
    SHEET1.merge_cells(start_row=crow+1, start_column=5,
                       end_row=crow+1, end_column=7)
    SHEET1.append(['Total', '', '', '', '', '', '', str(
        runs)+'('+str(wickets)+' wkts, '+ov[0]+' Ov)'])

    SHEET2.append([])
    SHEET2.append(['Powerplays', 'Overs', '', '', '', '', '', '', 'Runs'])
    SHEET2.append(['Mandotary', '0.1-6', '', '', '', '', '', '', ppr])

    #Combining the data from all the sheets.
    mc1 = SHEET1.max_column
    mc2 = SHEET2.max_column
    mc3 = SHEET3.max_column
    mr1 = SHEET1.max_row
    mr2 = SHEET2.max_row
    mr3 = SHEET3.max_row

    for i in range(1, mr1-3):
        SHEET1.merge_cells(start_row=i+1, start_column=2,end_row=i+1, end_column=5)
    for i in range(1, mr3+1):
        for j in range(1, 9):
            SHEET1.cell(row=mr1+i+1, column=j).value = SHEET3.cell(row=i, column=j).value

    mr11 = SHEET1.max_row
    for i in range(1, mr2+1):
        for j in range(1, 11):
            SHEET1.cell(row=mr11+i+3, column=j).value = SHEET2.cell(row=i, column=j).value

    #Some more formatting commands.
    SHEET1.merge_cells(start_row=mr1+2, start_column=1,end_row=mr1+2, end_column=10)
    SHEET1.merge_cells(start_row=mr1+3, start_column=1,end_row=mr1+5, end_column=10)
    SHEET1.cell(row=mr1+3, column=1).alignment = Alignment(wrap_text=True)
    mr12 = SHEET1.max_row
    SHEET1.merge_cells(start_row=mr12 - 1, start_column=2,end_row=mr12-1, end_column=8)
    SHEET1.merge_cells(start_row=mr12, start_column=2,end_row=mr12, end_column=8)
    SHEET1.merge_cells(start_row=mr12 - 1, start_column=9,end_row=mr12-1, end_column=10)
    SHEET1.merge_cells(start_row=mr12, start_column=9,end_row=mr12, end_column=10)

    return innbat

#This function will generate scorecard.
def scorecard(inn1,inn2):
    pak_pl={'Babar':'Babar Azam (c)','Rizwan':'Mohammad Rizwan (wk)','Fakhar':'Fakhar Zaman','Iftikhar':'Iftikhar Ahmed','Khushdil':'Khushdil Shah','Asif':'Asif Ali','Shadab':'Shadab Khan','Mohammad':'Mohammad Nawaz','Naseem':'Naseem Shah','Haris':'Haris Rauf','Dahani':'Shahnawaz Dahani'}
    ind_pl={'Rohit':'Rohit Sharma (c)','Rahul':'KL Rahul','Kohli':'Virat Kohli','Suryakumar':'Suryakumar Yadav','Jadeja':'Ravindra Jadeja','Hardik':'Hardik Pandya','Karthik':'Dinesh Karthik (wk)','Bhuvneshwar':'Bhuvneshwar Kumar','Avesh':'Avesh Khan',"Arshdeep":'Arshdeep Singh','Chahal':'Yuzvendra Chahal'}
    finn=INNING(inn1,pak_pl,ind_pl,'Pakistan')
    sinn=INNING(inn2,ind_pl,pak_pl,'India')

    SHEET1=finn.active
    SHEET2=sinn.active

    #Here we are combining both the innings.
    mc1=SHEET1.max_column
    mc2=SHEET2.max_column
    mr1=SHEET1.max_row
    mr2=SHEET2.max_row

    for i in range(1,mr2+1):
        for j in range(1,11):
            SHEET1.cell(row=mr1+i+1,column=j).value=SHEET2.cell(row=i,column=j).value    
    SHEET1.merge_cells(start_row=34, start_column=1, end_row=34, end_column=8)
    SHEET1.merge_cells(start_row=35, start_column=2, end_row=35, end_column=5)
    SHEET1.merge_cells(start_row=48, start_column=1, end_row=50, end_column=10)
    for i in range(36,43):
        SHEET1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    return finn
    
#Oppening the comentary files given in .yxt format.
pak=open("pak_inns1.txt",'r')
ind=open('india_inns2.txt','r')
team=open('teams.txt','r')

wb=Workbook()
wb=scorecard(pak,ind)

#Saving the output in .xlsx format.
wb.save('Scorecard.xlsx')
end_time=datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))

